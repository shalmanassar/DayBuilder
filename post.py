"""DayBuilder — post.py
Flatten blocks to TimeLogTable rows, write to target workbook, backup, sync."""
import os
import json
import shutil
import time as _time
from datetime import datetime, date
from pathlib import Path

import db

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


BASE_DIR = Path(__file__).parent.resolve()

# Day-of-week → column letter (1-indexed: B=2, C=3, ...)
DAY_COLUMNS = {0: 'B', 1: 'C', 2: 'D', 3: 'E', 4: 'F'}  # Mon=0..Fri=4

# Day-of-week → comment row
DAY_COMMENT_ROWS = {0: 26, 1: 28, 2: 30, 3: 32, 4: 34}


def validate_blocks(blocks):
    """Pre-post validation. Returns {hard: [...], soft: [...]}."""
    hard = []
    soft = []

    types = [b.get('type') for b in blocks]
    if 'clock_in' not in types:
        hard.append('Clock in time not set')
    if 'clock_out' not in types:
        hard.append('Clock out time not set')

    # Must have at least one work block (not just template placeholders)
    work_types = [t for t in types if t not in ('clock_in', 'clock_out', 'break', 'lunch')]
    if not work_types:
        hard.append('No work activities logged — add at least one activity before posting')

    # Check all times explicit
    for b in blocks:
        if not b.get('start') or (b.get('type') not in ('clock_in', 'clock_out') and not b.get('end')):
            hard.append(f"Block '{b.get('type','?')}' missing explicit time")
            break

    # Breaks
    breaks = [b for b in blocks if b.get('type') == 'break']
    break_mins = sum(_duration(b) for b in breaks)
    if len(breaks) < 2:
        soft.append(f'Only {len(breaks)} break(s) — expected 2')
    elif any(_duration(b) < 15 for b in breaks):
        soft.append('A break is shorter than 15 minutes')

    # Lunch
    lunches = [b for b in blocks if b.get('type') == 'lunch']
    if not lunches:
        soft.append('No lunch block')
    elif any(_duration(b) < 30 for b in lunches):
        soft.append('Lunch is shorter than 30 minutes')

    # Gaps — unaccounted time blocks posting
    sorted_blocks = sorted([b for b in blocks if b.get('start') and b.get('end')],
                           key=lambda b: b['start'])
    for i in range(1, len(sorted_blocks)):
        if sorted_blocks[i]['start'] > sorted_blocks[i-1]['end']:
            hard.append(f"Unaccounted time: {sorted_blocks[i-1]['end']} – {sorted_blocks[i]['start']}")

    return {'hard': hard, 'soft': soft}


def flatten_blocks(blocks, date_iso, user_id):
    """Convert DayDraft blocks → TimeLogTable rows (master DB format: JDN + unix ts)."""
    jdn = db.iso_to_jdn(date_iso)
    ts = int(_time.time())
    rows = []

    for seq, block in enumerate(sorted(blocks, key=lambda b: b.get('start', ''))):
        btype = block.get('type', '')
        job = btype if btype in ('clock_in', 'clock_out') else btype

        start_time = block.get('start', '')
        time_unix = db.time_to_unix(start_time, date_iso) if start_time else ts
        elapsed = _elapsed_str(block) if block.get('start') and block.get('end') else None

        memo_parts = []
        if block.get('subtype'):
            memo_parts.append(block['subtype'])
        if block.get('memo'):
            memo_parts.append(block['memo'])
        memo = ' - '.join(memo_parts) if memo_parts else ''

        rows.append({
            'uid': f"{user_id}_{seq}_{ts}",
            'date': jdn,
            'job': job,
            'time': time_unix,
            'e_time': elapsed,
            'memo': memo,
            'device': block.get('device') or '',
            'qty': block.get('qty') or ''
        })

    return rows


def aggregate_productivity(blocks, shared_config):
    """Aggregate device quantities and event counts from blocks."""
    device_types = {d['id']: d for d in shared_config.get('device_types', [])}
    asset_paths = {p['id']: p for p in shared_config.get('asset_paths', [])}
    counts = {d['id']: 0 for d in shared_config.get('device_types', [])}

    for block in blocks:
        # Device qty
        if block.get('device') and block.get('qty'):
            dev = block['device']
            if dev in counts:
                counts[dev] += block['qty']

        # Event counts (RTV_Events, Prov_Events)
        if block.get('subtype') and block['subtype'] in asset_paths:
            path_info = asset_paths[block['subtype']]
            for event_type in path_info.get('counts_toward', []):
                if event_type in counts:
                    counts[event_type] += 1

    return counts


def calculate_hours(blocks):
    """Calculate working and non-working hours."""
    work_mins = 0
    nonwork_mins = 0
    for b in blocks:
        dur = _duration(b)
        if b.get('type') in ('break', 'lunch'):
            nonwork_mins += dur
        elif b.get('type') not in ('clock_in', 'clock_out'):
            work_mins += dur
    return round(work_mins / 60, 2), round(nonwork_mins / 60, 2)


def calculate_rates(blocks, shared_config):
    """Legacy wrapper — returns old format for backward compat."""
    report = calculate_report(blocks, shared_config)
    return report.get('devices', [])


def calculate_report(blocks, shared_config, schedule=None):
    """Full productivity report using quota-hours model.

    Model: qty / quota_per_hour = equivalent production hours for that device.
    Sum all device hours = total production achieved.
    Compare to available production hours (shift minus off-clock time).

    Returns dict with devices, off_clock, non_asset, totals.
    """
    quotas = shared_config.get('quotas', {})
    device_types = {d['id']: d['display'] for d in shared_config.get('device_types', [])}

    # Schedule defaults
    if not schedule:
        schedule = {}
    sched_start = schedule.get('default_start', '08:00')
    sched_end = schedule.get('default_end', '16:30')
    sched_break_count = schedule.get('break_count', 2)
    sched_break_min = schedule.get('break_minutes', 15)
    sched_lunch_min = schedule.get('lunch_minutes', 30)

    # Scheduled shift and off-clock
    shift_mins = _time_to_min(sched_end) - _time_to_min(sched_start)
    scheduled_off = sched_lunch_min + (sched_break_count * sched_break_min)
    scheduled_prod_mins = shift_mins - scheduled_off

    # Actual clock in/out
    clock_in_block = next((b for b in blocks if b.get('type') == 'clock_in'), None)
    clock_out_block = next((b for b in blocks if b.get('type') == 'clock_out'), None)
    actual_in = clock_in_block.get('start') if clock_in_block else sched_start
    actual_out = clock_out_block.get('start') if clock_out_block else sched_end

    # Off-clock time detection
    off_clock = []
    late_in_mins = max(0, _time_to_min(actual_in) - _time_to_min(sched_start))
    if late_in_mins > 0:
        off_clock.append({'label': 'Late clock-in', 'minutes': late_in_mins,
                          'detail': f"{actual_in} vs scheduled {sched_start}"})

    early_out_mins = max(0, _time_to_min(sched_end) - _time_to_min(actual_out))
    if early_out_mins > 0:
        off_clock.append({'label': 'Early clock-out', 'minutes': early_out_mins,
                          'detail': f"{actual_out} vs scheduled {sched_end}"})

    # Actual break/lunch durations vs scheduled
    actual_break_mins = sum(_duration(b) for b in blocks if b.get('type') == 'break')
    actual_lunch_mins = sum(_duration(b) for b in blocks if b.get('type') == 'lunch')
    expected_break_mins = sched_break_count * sched_break_min

    excess_break = max(0, actual_break_mins - expected_break_mins)
    if excess_break > 0:
        off_clock.append({'label': 'Excess break time', 'minutes': excess_break,
                          'detail': f"{actual_break_mins}m actual vs {expected_break_mins}m scheduled"})

    excess_lunch = max(0, actual_lunch_mins - sched_lunch_min)
    if excess_lunch > 0:
        off_clock.append({'label': 'Excess lunch time', 'minutes': excess_lunch,
                          'detail': f"{actual_lunch_mins}m actual vs {sched_lunch_min}m scheduled"})

    total_off_clock_excess = late_in_mins + early_out_mins + excess_break + excess_lunch

    # Available production hours = scheduled production - off-clock excess
    available_prod_mins = max(0, scheduled_prod_mins - total_off_clock_excess)
    available_prod_hrs = available_prod_mins / 60

    # Aggregate qty per device and compute quota-hours
    counts = aggregate_productivity(blocks, shared_config)
    devices = []
    total_quota_hrs = 0

    for dev_id, qty in counts.items():
        if qty <= 0:
            continue
        quota = quotas.get(dev_id, 0)
        quota_hrs = (qty / quota) if quota > 0 else 0
        pct_of_day = (quota_hrs / available_prod_hrs * 100) if available_prod_hrs > 0 else 0
        total_quota_hrs += quota_hrs
        devices.append({
            'device': dev_id,
            'display': device_types.get(dev_id, dev_id),
            'qty': qty,
            'quota': quota,
            'quota_hrs': round(quota_hrs, 2),
            'pct_of_day': round(pct_of_day, 1)
        })

    # Overall production percentage (against full available time)
    overall_pct = (total_quota_hrs / available_prod_hrs * 100) if available_prod_hrs > 0 else 0

    # Non-asset time breakdown — individual blocks with detail
    non_asset_types = ('project', 'admin', 'meeting', '5s', 'learning')
    non_asset = []
    total_non_asset_mins = 0
    for b in blocks:
        if b.get('type') in non_asset_types:
            dur = _duration(b)
            if dur > 0:
                non_asset.append({
                    'type': b['type'],
                    'label': b['type'].replace('_', ' ').title(),
                    'memo': b.get('memo') or '',
                    'start': b.get('start'),
                    'end': b.get('end'),
                    'minutes': dur,
                    'hours': round(dur / 60, 2)
                })
                total_non_asset_mins += dur

    # Adjusted production goal: deduct non-asset time from available
    adjusted_prod_mins = max(0, available_prod_mins - total_non_asset_mins)
    adjusted_prod_hrs = adjusted_prod_mins / 60
    adjusted_pct = (total_quota_hrs / adjusted_prod_hrs * 100) if adjusted_prod_hrs > 0 else 0

    # Synopsis
    synopsis = _build_synopsis(blocks, devices, non_asset, off_clock, total_quota_hrs,
                               adjusted_prod_hrs, adjusted_pct, actual_in, actual_out, sched_start, sched_end)

    return {
        'devices': devices,
        'off_clock': off_clock,
        'non_asset': non_asset,
        'synopsis': synopsis,
        'totals': {
            'shift_hours': round(shift_mins / 60, 2),
            'scheduled_prod_hours': round(scheduled_prod_mins / 60, 2),
            'off_clock_excess_mins': total_off_clock_excess,
            'available_prod_hours': round(available_prod_hrs, 2),
            'total_quota_hours': round(total_quota_hrs, 2),
            'overall_pct': round(overall_pct, 1),
            'adjusted_prod_hours': round(adjusted_prod_hrs, 2),
            'adjusted_pct': round(adjusted_pct, 1),
            'non_asset_hours': round(total_non_asset_mins / 60, 2),
            'non_asset_pct': round((total_non_asset_mins / 60 / available_prod_hrs * 100) if available_prod_hrs > 0 else 0, 1),
            'actual_in': actual_in,
            'actual_out': actual_out,
            'actual_break_mins': actual_break_mins,
            'actual_lunch_mins': actual_lunch_mins
        }
    }


def _build_synopsis(blocks, devices, non_asset, off_clock, total_quota_hrs,
                    adjusted_prod_hrs, adjusted_pct, actual_in, actual_out, sched_start, sched_end):
    """Generate a human-readable daily synopsis from block data."""
    parts = []

    # Clock in/out note
    if actual_in != sched_start:
        diff = _time_to_min(actual_in) - _time_to_min(sched_start)
        if diff > 0:
            parts.append(f"Clocked in at {actual_in} ({diff}m late)")
        else:
            parts.append(f"Clocked in at {actual_in} ({-diff}m early)")
    else:
        parts.append(f"Clocked in on time at {actual_in}")

    # Device production
    if devices:
        dev_parts = [f"{d['qty']} {d['display']}" for d in devices]
        parts.append("Processed " + ", ".join(dev_parts))

    # Non-asset activities
    if non_asset:
        for na in non_asset:
            memo = f" ({na['memo']})" if na['memo'] else ''
            parts.append(f"{na['minutes']}m {na['label']}{memo}")

    # Off-clock notes
    for oc in off_clock:
        if 'Excess' in oc['label']:
            parts.append(f"{oc['minutes']}m {oc['label'].lower()}")

    # Result
    parts.append(f"Achieved {round(adjusted_pct, 1)}% adjusted production goal ({round(total_quota_hrs, 2)}h of {round(adjusted_prod_hrs, 2)}h)")

    if actual_out != sched_end:
        diff = _time_to_min(sched_end) - _time_to_min(actual_out)
        if diff > 0:
            parts.append(f"Clocked out at {actual_out} ({diff}m early)")
        else:
            parts.append(f"Clocked out at {actual_out} ({-diff}m late)")

    return ". ".join(parts) + "."


def build_comment(blocks):
    """Build day comment from block memos."""
    parts = []
    for b in blocks:
        if b.get('type') in ('clock_in', 'clock_out', 'break', 'lunch'):
            continue
        label = b.get('type', '').replace('_', ' ').title()
        detail = b.get('memo') or ''
        if b.get('device'):
            detail = f"{b['device']} x{b.get('qty','')} {detail}".strip()
        if detail:
            parts.append(f"{label}: {detail}")
    return '; '.join(parts)


def backup_workbook(target_path):
    """Copy target workbook to backup/ with timestamp."""
    if not os.path.isfile(target_path):
        return None
    backup_dir = BASE_DIR / "backup"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = Path(target_path).stem
    dest = backup_dir / f"{name}_{ts}.xlsm"
    shutil.copy2(target_path, dest)
    return str(dest)


def write_workbook(target_path, sheet_name, date_iso, blocks, shared_config):
    """Write productivity data to target workbook. Returns None on success, error string on failure."""
    if load_workbook is None:
        return "openpyxl not installed"

    # Determine column from day of week
    d = date.fromisoformat(date_iso)
    weekday = d.weekday()  # 0=Mon
    if weekday not in DAY_COLUMNS:
        return f"Date {date_iso} is not a weekday (Mon-Fri)"

    col_letter = DAY_COLUMNS[weekday]
    comment_row = DAY_COMMENT_ROWS[weekday]

    # Try to open workbook with retries
    wb = None
    for attempt in range(3):
        try:
            wb = load_workbook(target_path, keep_vba=True)
            break
        except Exception as e:
            if attempt < 2:
                _time.sleep(2)
            else:
                return f"Cannot open workbook after 3 attempts: {e}"

    if sheet_name not in wb.sheetnames:
        return f"Sheet '{sheet_name}' not found in workbook"

    ws = wb[sheet_name]

    # Write productivity (rows 7-19) — only write non-zero values
    counts = aggregate_productivity(blocks, shared_config)
    for dt in shared_config.get('device_types', []):
        row = dt['row']
        qty = counts.get(dt['id'], 0)
        ws[f"{col_letter}{row}"] = qty

    # Write hours (rows 22-23)
    work_hours, nonwork_hours = calculate_hours(blocks)
    ws[f"{col_letter}22"] = work_hours
    ws[f"{col_letter}23"] = nonwork_hours

    # Write comment (use report synopsis)
    report = calculate_report(blocks, shared_config)
    comment = report.get('synopsis', '') or build_comment(blocks)
    ws[f"B{comment_row}"] = comment

    # Clear any accidental fill on rows 19-66 (openpyxl can corrupt formatting)
    from openpyxl.styles import PatternFill
    no_fill = PatternFill(fill_type=None)
    for row in range(19, 67):
        for col in ('B', 'C', 'D', 'E', 'F'):
            cell = ws[f"{col}{row}"]
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                cell.fill = no_fill

    try:
        wb.save(target_path)
    except Exception as e:
        return f"Failed to save workbook: {e}"

    return None


def sync_db(db_path, sync_target, user_id):
    """Copy local DB to shared sync target. Best-effort."""
    if not sync_target or not os.path.isdir(sync_target):
        return False
    try:
        dest = os.path.join(sync_target, f"{user_id}_timelog.db")
        shutil.copy2(db_path, dest)
        return True
    except Exception:
        return False


def post_day(date_iso, cfg, shared_config, db_path):
    """Full post operation. Returns {ok, error, warnings, backup_path}."""
    draft = db.get_draft(date_iso, db_path)
    if not draft or not draft['blocks']:
        return {'ok': False, 'error': 'No blocks to post'}

    blocks = draft['blocks']

    # 1. Validate
    validation = validate_blocks(blocks)
    if validation['hard']:
        return {'ok': False, 'error': 'Validation failed', 'hard': validation['hard'], 'soft': validation['soft']}

    # 2. Flatten
    user_id = cfg.get('user_id', 'unknown')
    rows = flatten_blocks(blocks, date_iso, user_id)

    # 3. Backup
    target_path = cfg.get('target_workbook')
    backup_path = None
    if target_path and os.path.isfile(target_path):
        backup_path = backup_workbook(target_path)

    # 4. Write workbook
    wb_error = None
    if target_path:
        sheet_name = cfg.get('target_sheet', cfg.get('user_display_name', 'Sheet1'))
        wb_error = write_workbook(target_path, sheet_name, date_iso, blocks, shared_config)

    if wb_error:
        return {'ok': False, 'error': wb_error, 'soft': validation['soft'], 'backup_path': backup_path}

    # 5. Commit to SQLite
    try:
        db.delete_timelog_by_date(date_iso, db_path)
        db.insert_timelog_rows(rows, db_path)
        # Mark draft as posted
        conn = db.get_db(db_path)
        conn.execute("UPDATE DayDraft SET posted = 1, posted_at = ? WHERE date = ?",
                     (datetime.now().isoformat(), date_iso))
        conn.commit()
        conn.close()
    except Exception as e:
        return {'ok': True, 'warning': f'Posted to workbook but DB update failed: {e}',
                'soft': validation['soft'], 'backup_path': backup_path}

    # 6. Sync
    sync_ok = sync_db(db_path, cfg.get('sync_target'), user_id)

    return {'ok': True, 'soft': validation['soft'], 'backup_path': backup_path, 'synced': sync_ok}


# --- Helpers ---

def _duration(block):
    """Duration in minutes."""
    s, e = block.get('start'), block.get('end')
    if not s or not e:
        return 0
    return _time_to_min(e) - _time_to_min(s)


def _time_to_min(t):
    parts = t.split(':')
    return int(parts[0]) * 60 + int(parts[1])


def _to_12h(t24):
    """HH:MM → H:MM:SS AM/PM"""
    h, m = int(t24.split(':')[0]), int(t24.split(':')[1])
    ampm = 'AM' if h < 12 else 'PM'
    h12 = h % 12 or 12
    return f"{h12}:{m:02d}:00 {ampm}"


def _elapsed_str(block):
    """Calculate elapsed time string H:MM:SS."""
    mins = _duration(block)
    h = mins // 60
    m = mins % 60
    return f"{h}:{m:02d}:00"
