from openpyxl import load_workbook
wb = load_workbook(r'd:\DayBuilder\manager\Total Report_reference.xlsx', data_only=False)

with open(r'd:\DayBuilder\manager\report_dump.txt', 'w', encoding='utf-8') as f:
    for sn in wb.sheetnames:
        ws = wb[sn]
        f.write(f'=== {sn} (rows={ws.max_row}, cols={ws.max_column}) ===\n')
        for row in range(1, min(70, ws.max_row+1)):
            vals = []
            for col in range(1, min(20, ws.max_column+1)):
                c = ws.cell(row=row, column=col)
                v = c.value
                if v is not None:
                    vals.append(f'{c.coordinate}={repr(v)[:80]}')
            if vals:
                f.write(f'  R{row}: {" | ".join(vals)}\n')
        f.write('\n')
    f.write('DONE\n')
