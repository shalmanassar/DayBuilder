"""Workbook importer — reads employee workbooks, imports to master DB."""
import os
import time
import shutil
import sqlite3
import logging
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl

import governance

logger = logging.getLogger("manager")

# Workbook layout constants
DEVICE_ROWS = range(7, 20)       # rows 7-19 inclusive
HOURS_ROW = 22
NON_WORK_ROW = 23
COMMENT_ROWS = {0: 26, 1: 28, 2: 30, 3: 32, 4: 34}  # weekday index -> row
DAY_COLS = {"B": 0, "C": 1, "D": 2, "E": 3, "F": 4}  # col -> weekday offset (Mon=0)
COL_INDICES = [2, 3, 4, 5, 6]  # B=2, C=3, D=4, E=5, F=6 (1-indexed openpyxl)


def _master_db_path(cfg):
    return Path(cfg["rma_job_logger_path"]) / "POST" / "m_timelog.db"


def _lock_path(cfg):
    return Path(cfg["rma_job_logger_path"]) / "POST" / "m_timelog.lock"


def _date_to_jdn(d):
    """Convert a date to Julian Day Number."""
    a = (14 - d.month) // 12
    y = d.year + 4800 - a
    m = d.month + 12 * a - 3
    return d.day + (153 * m + 2) // 5 + 365 * y + y // 4 - y // 100 + y // 400 - 32045


def _jdn_to_date(jdn):
    """Convert Julian Day Number back to date."""
    a = jdn + 32044
    b = (4 * a + 3) // 146097
    c = a - (146097 * b) // 4
    d_ = (4 * c + 3) // 1461
    e = c - (1461 * d_) // 4
    m = (5 * e + 2) // 153
    day = e - (153 * m + 2) // 5 + 1
    month = m + 3 - 12 * (m // 10)
    year = 100 * b + d_ - 4800 + m // 10
    return date(year, month, day)


def _get_device_map(cfg):
    """Build row -> system_id mapping from shared_config device_types."""
    sc = governance.load_shared_config(cfg)
    mapping = {}
    for dt in sc.get("device_types", []):
        row = dt.get("row")
        if row:
            mapping[row] = dt["id"]
    return mapping


def detect_week_monday(workbook_path):
    """Determine the Monday of the week the workbook data represents.
    Uses file modification time as hint — the week containing that date."""
    mtime = os.path.getmtime(workbook_path)
    mod_date = datetime.fromtimestamp(mtime).date()
    # Monday of that week (weekday: Mon=0)
    monday = mod_date - timedelta(days=mod_date.weekday())
    return monday


def read_workbook(workbook_path, sheet_name=None):
    """Read a single employee workbook. Returns dict of day data.
    
    Returns: {
        weekday_index(0-4): {
            "devices": {system_id: qty, ...},
            "hours": float,
            "non_work_hours": float,
            "comment": str
        }, ...
    }
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    days = {}
    for col_idx, weekday in zip(COL_INDICES, range(5)):
        devices = {}
        has_data = False
        for row in DEVICE_ROWS:
            val = ws.cell(row=row, column=col_idx).value
            if val and isinstance(val, (int, float)) and val > 0:
                devices[row] = int(val)
                has_data = True
        
        hours = ws.cell(row=HOURS_ROW, column=col_idx).value or 0
        non_work = ws.cell(row=NON_WORK_ROW, column=col_idx).value or 0
        comment_row = COMMENT_ROWS.get(weekday)
        comment = ""
        if comment_row:
            comment = ws.cell(row=comment_row, column=2).value or ""  # column B
        
        if has_data or hours:
            days[weekday] = {
                "devices": devices,
                "hours": float(hours) if hours else 0,
                "non_work_hours": float(non_work) if non_work else 0,
                "comment": str(comment).strip()
            }
    
    wb.close()
    return days


def acquire_lock(cfg, timeout=10):
    """Acquire master DB lock file. Returns True if acquired."""
    lock = _lock_path(cfg)
    lock.parent.mkdir(parents=True, exist_ok=True)
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            fd = os.open(str(lock), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode())
            os.close(fd)
            return True
        except FileExistsError:
            time.sleep(0.2)
    return False


def release_lock(cfg):
    """Release master DB lock file."""
    lock = _lock_path(cfg)
    try:
        os.remove(lock)
    except OSError:
        pass


def backup_master_db(cfg):
    """Create timestamped backup of master DB before import."""
    db = _master_db_path(cfg)
    if db.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = db.with_name(f"m_timelog_backup_{ts}.db")
        shutil.copy2(db, backup)
        return str(backup)
    return None


def _ensure_master_db(cfg):
    """Create master DB and table if not exists."""
    db = _master_db_path(cfg)
    db.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db))
    conn.execute("""CREATE TABLE IF NOT EXISTS TimeLogTable (
        uid TEXT PRIMARY KEY,
        date DATE,
        job TEXT,
        time TIME,
        e_time TIME,
        memo TEXT,
        device TEXT,
        qty INT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS ImportMeta (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT,
        week_of TEXT,
        imported_at TEXT,
        rows_inserted INT
    )""")
    conn.commit()
    return conn


def get_existing_entries(cfg, user_id, monday):
    """Check what data already exists in master DB for user+week."""
    db = _master_db_path(cfg)
    if not db.exists():
        return {}
    conn = sqlite3.connect(str(db))
    # Get JDNs for Mon-Fri
    existing = {}
    for offset in range(5):
        day = monday + timedelta(days=offset)
        jdn = _date_to_jdn(day)
        rows = conn.execute(
            "SELECT device, qty FROM TimeLogTable WHERE uid LIKE ? AND date=? AND uid NOT LIKE 'v-%'",
            (f"{user_id}_%", jdn)
        ).fetchall()
        if rows:
            existing[offset] = {r[0]: r[1] for r in rows}
    conn.close()
    return existing


def import_employee(cfg, employee, device_map=None):
    """Import a single employee's workbook into master DB.
    
    Returns: {
        "user_id": str,
        "status": "ok"|"skipped"|"error",
        "imported": int,  # days imported
        "skipped": int,   # days already in DB
        "discrepancies": [{day, device, db_val, wb_val}, ...],
        "error": str|None
    }
    """
    user_id = employee["user_id"]
    wb_path = employee.get("target_workbook", "")
    sheet = employee.get("target_sheet")
    
    if not wb_path or not os.path.isfile(wb_path):
        return {"user_id": user_id, "status": "skipped", "imported": 0,
                "skipped": 0, "discrepancies": [], "error": "workbook not found"}
    
    if device_map is None:
        device_map = _get_device_map(cfg)
    
    try:
        monday = detect_week_monday(wb_path)
        days = read_workbook(wb_path, sheet)
    except Exception as e:
        return {"user_id": user_id, "status": "error", "imported": 0,
                "skipped": 0, "discrepancies": [], "error": str(e)}
    
    if not days:
        return {"user_id": user_id, "status": "skipped", "imported": 0,
                "skipped": 0, "discrepancies": [], "error": "no data in workbook"}
    
    existing = get_existing_entries(cfg, user_id, monday)
    
    imported = 0
    skipped = 0
    discrepancies = []
    rows_to_insert = []
    import_ts = int(time.time())
    
    for weekday, day_data in days.items():
        day_date = monday + timedelta(days=weekday)
        jdn = _date_to_jdn(day_date)
        
        if weekday in existing:
            # Check for discrepancies
            for row_num, qty in day_data["devices"].items():
                device_id = device_map.get(row_num, f"row_{row_num}")
                db_qty = existing[weekday].get(device_id, 0)
                if db_qty != qty:
                    discrepancies.append({
                        "date": day_date.isoformat(),
                        "device": device_id,
                        "db_value": db_qty,
                        "wb_value": qty
                    })
            skipped += 1
            continue
        
        # Generate insert rows
        for row_num, qty in day_data["devices"].items():
            device_id = device_map.get(row_num, f"row_{row_num}")
            uid = f"{user_id}_imp_{jdn}_{row_num}_{import_ts}"
            memo = day_data.get("comment", "")
            rows_to_insert.append((uid, jdn, "RMA_PTS", import_ts, "", memo, device_id, qty))
        imported += 1
    
    # Write to DB if we have rows
    if rows_to_insert:
        if not acquire_lock(cfg):
            return {"user_id": user_id, "status": "error", "imported": 0,
                    "skipped": skipped, "discrepancies": discrepancies, "error": "could not acquire lock"}
        try:
            backup_master_db(cfg)
            conn = _ensure_master_db(cfg)
            conn.executemany(
                "INSERT OR IGNORE INTO TimeLogTable (uid, date, job, time, e_time, memo, device, qty) VALUES (?,?,?,?,?,?,?,?)",
                rows_to_insert
            )
            conn.execute(
                "INSERT INTO ImportMeta (user_id, week_of, imported_at, rows_inserted) VALUES (?,?,?,?)",
                (user_id, monday.isoformat(), datetime.now().isoformat(), len(rows_to_insert))
            )
            conn.commit()
            conn.close()
        finally:
            release_lock(cfg)
    
    status = "ok" if imported > 0 else "skipped"
    return {"user_id": user_id, "status": status, "imported": imported,
            "skipped": skipped, "discrepancies": discrepancies, "error": None}


def import_all(cfg):
    """Import all active employees' workbooks. Returns summary."""
    employees_data = governance.load_employees(cfg)
    employees = [e for e in employees_data.get("employees", []) if e.get("active", True)]
    device_map = _get_device_map(cfg)
    
    results = []
    for emp in employees:
        result = import_employee(cfg, emp, device_map)
        results.append(result)
        logger.info(f"Import {emp['user_id']}: {result['status']} "
                    f"({result['imported']} imported, {result['skipped']} skipped)")
    
    total_imported = sum(r["imported"] for r in results)
    total_skipped = sum(r["skipped"] for r in results)
    all_discrepancies = []
    for r in results:
        for d in r["discrepancies"]:
            d["user_id"] = r["user_id"]
            all_discrepancies.append(d)
    
    return {
        "timestamp": datetime.now().isoformat(),
        "employees_scanned": len(employees),
        "total_imported": total_imported,
        "total_skipped": total_skipped,
        "discrepancies": all_discrepancies,
        "results": results
    }


def resolve_discrepancy(cfg, user_id, date_str, device_id, action="keep_db"):
    """Resolve a discrepancy. action='accept_workbook' voids old + inserts new."""
    if action == "keep_db":
        return {"ok": True, "action": "kept DB value"}
    
    # accept_workbook: void existing entry, re-import from workbook
    day_date = date.fromisoformat(date_str)
    jdn = _date_to_jdn(day_date)
    
    if not acquire_lock(cfg):
        return {"ok": False, "error": "could not acquire lock"}
    
    try:
        conn = _ensure_master_db(cfg)
        # Find and void existing
        rows = conn.execute(
            "SELECT uid FROM TimeLogTable WHERE uid LIKE ? AND date=? AND device=? AND uid NOT LIKE 'v-%'",
            (f"{user_id}_%", jdn, device_id)
        ).fetchall()
        for (uid,) in rows:
            conn.execute("UPDATE TimeLogTable SET uid=? WHERE uid=?", (f"v-{uid}", uid))
        
        # Get fresh value from workbook
        emp = governance.get_employee(cfg, user_id)
        if emp and emp.get("target_workbook") and os.path.isfile(emp["target_workbook"]):
            monday = detect_week_monday(emp["target_workbook"])
            weekday = (day_date - monday).days
            if 0 <= weekday <= 4:
                days = read_workbook(emp["target_workbook"], emp.get("target_sheet"))
                if weekday in days:
                    device_map = _get_device_map(cfg)
                    for row_num, qty in days[weekday]["devices"].items():
                        if device_map.get(row_num) == device_id:
                            uid = f"{user_id}_imp_{jdn}_{row_num}_{int(time.time())}"
                            conn.execute(
                                "INSERT INTO TimeLogTable (uid,date,job,time,e_time,memo,device,qty) VALUES (?,?,?,?,?,?,?,?)",
                                (uid, jdn, "RMA_PTS", int(time.time()), "", "", device_id, qty)
                            )
        conn.commit()
        conn.close()
    finally:
        release_lock(cfg)
    
    return {"ok": True, "action": "accepted workbook value"}
